import openpyxl
from datetime import datetime
from copy import deepcopy

def add_sheet_to_excel(sheet_name, purchase_date,template_file_path):
    try:
        # Load the existing workbook
        wb = openpyxl.load_workbook(filename=template_file_path)
        
        # Get the sheet to duplicate
        sheet_to_duplicate = wb.worksheets[0]
        
        # Create a new sheet as a copy of the original sheet
        new_sheet = wb.copy_worksheet(sheet_to_duplicate)
        
        # Rename the new sheet
        new_sheet.title = sheet_name
        
        # Duplicate all charts from the original sheet to the new sheet
        for chart in sheet_to_duplicate._charts:
            new_chart = deepcopy(chart)
            new_sheet.add_chart(new_chart)

        # Parse the purchase date
        date_obj = datetime.strptime(purchase_date, "%m/%d/%Y")
        
        # Set the purchase date in the specified cells
        new_sheet['C2'] = date_obj
        new_sheet['C8'] = date_obj
        new_sheet['C14'] = date_obj
        
                # Apply the desired date format
        date_format = 'mmm-yy'
        new_sheet['C2'].number_format = date_format
        new_sheet['C8'].number_format = date_format
        new_sheet['C14'].number_format = date_format
        
        # Save the changes
        wb.save(filename=template_file_path)
        
        print(f"Sheet '{sheet_name}' duplicated and renamed to '{sheet_name}' in '{template_file_path}'.")
        return True
    except Exception as e:
        print(f"Error occurred: {e}")
        return False

def remove_sheet_from_excel(sheet_name, template_file_path):
    try:
        # Load the Excel template
        wb = openpyxl.load_workbook(template_file_path)
        
        # Check if the sheet exists in the workbook
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Remove the sheet
            wb.remove(ws)
            
            # Save the changes to the template file
            wb.save(template_file_path)
            print(f"Sheet '{sheet_name}' removed from Excel template.")
            return True
        else:
            print(f"Sheet '{sheet_name}' not found in Excel template.")
            return True
    
    except Exception as e:
        print(f"Error occurred while removing sheet '{sheet_name}': {e}")
        return False


def load_sheet_names(file_path, template_file_path):
    try:
        # Load the template workbook
        wb = openpyxl.load_workbook(filename=template_file_path)
        # Get the sheet names
        sheet_names = [sheet.title for sheet in wb.worksheets]
        
        # Write the sheet names to the file
        with open(file_path, 'w') as file:
            file.write('\n'.join(sheet_names))
        
        return sheet_names
    except FileNotFoundError:
        return []
